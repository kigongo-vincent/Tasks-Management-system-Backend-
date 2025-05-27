from django.shortcuts import render
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework.decorators import api_view
from .serializers import *
from rest_framework.response import Response
from rest_framework import status
from django.contrib.auth.hashers import make_password
from django.db.models import Q
from datetime import datetime
from .utils import generate_task_change_action
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.db.models import Sum


class CustomTokenObtainPairSerializer(TokenObtainPairSerializer):

    @classmethod

    def get_token(cls, user):

        token = super().get_token(user)

        # Add custom claims to the token payload

        token['role'] = user.role

        token['username'] = user.username

        token['email'] = user.email

        token['contact'] = user.contact

        try:
            token['company'] = user.department.company.id
        except:
            token['company'] = None

        return token


# creating a custom serializer for information encoded in the tokens

class CustomTokenObtainPairView(TokenObtainPairView):

    serializer_class = CustomTokenObtainPairSerializer


@api_view(['POST', 'GET'])
def companies(request):
    try:
        if request.method == "POST":
            serialized = CompanySerializer(data = request.data)
            if serialized.is_valid():
                serialized.save()
                serialized.data["admin_email"] = UserSerializer(User.objects.get(id = serialized.data["admin"])).data["email"]
                return Response(serialized.data, status=status.HTTP_201_CREATED)
            else:
                return Response(status=status.HTTP_400_BAD_REQUEST)

        companies = Company.objects.all()
        serialized = CompanySerializer(companies, many = True)
        for company in serialized.data:
            company["admin_email"] = UserSerializer(User.objects.get(id = company["admin"])).data["email"]
        return Response(serialized.data)
    except:
        return Response(status=status.HTTP_418_IM_A_TEAPOT)

@api_view(['PATCH', 'DELETE', 'GET'])
def company(request, pk):

    try:
        company =Company.objects.get(id = pk)
        company_head = User.objects.get(id = company.admin.id)

        if request.method == "PATCH":
            print(request.data)

            serialized = CompanySerializer(company, data = request.data, partial = True)
            if serialized.is_valid():
                serialized.save()
                return Response(serialized.data, status=status.HTTP_202_ACCEPTED)
            else:
                return Response(status=status.HTTP_400_BAD_REQUEST)

        elif request.method == "DELETE":
            try:
                company_head = User.objects.get(id = company.admin.id)
                company_head.delete()
            except:
                pass
            company.delete()
            return Response(status=status.HTTP_202_ACCEPTED)
        elif request.method == "GET":
            serialized = CompanySerializer(company)
            return Response(serialized.data)
    except:
        return Response(status=status.HTTP_418_IM_A_TEAPOT)

@api_view(['PATCH', 'DELETE', 'GET'])
def department(request, pk):

    try:
        department =Department.objects.get(id = pk)

        if request.method == "PATCH":
            serialized = DepartmentSerializer(department, data = request.data, partial = True)
            if serialized.is_valid():
                serialized.save()
                return Response(serialized.data, status=status.HTTP_202_ACCEPTED)
            else:
                return Response(status=status.HTTP_400_BAD_REQUEST)

        elif request.method == "DELETE":
            try:
                department_head = User.objects.get(id = department.admin.id)
                department_head.delete()
            except:
                pass
            department.delete()
            return Response(status=status.HTTP_202_ACCEPTED)
        elif request.method == "GET":
            serialized = DepartmentSerializer(department)
            return Response(serialized.data)
    except:
        return Response(status=status.HTTP_418_IM_A_TEAPOT)

@api_view(['PATCH', 'DELETE', 'GET'])
def task(request, pk):
    try:
        task = Task.objects.get(id=pk)

        if request.method == "PATCH":
            # Get old task data before saving the update
            old_task = TaskSerializer(task).data

            # Serialize the request data (partial update)
            serialized = TaskSerializer(task, data=request.data, partial=True)

            if serialized.is_valid():
                # Save the changes to the task in the database
                serialized.save()

                # Reload the updated task from the database to get the latest data
                task.refresh_from_db()
                new_task = TaskSerializer(task).data

                # Only create logs for non-drafted tasks
                if not task.drafted:
                    # Print the changes to debug
                    print(generate_task_change_action(old_task, new_task))

                    # Log the change action
                    Logger.objects.create(
                        action=generate_task_change_action(old_task, new_task),
                        user=User.objects.get(id=task.employee.id)
                    )

                print(f"old: {old_task}\nnew: {new_task}")
                return Response(new_task, status=status.HTTP_202_ACCEPTED)
            else:
                return Response(status=status.HTTP_400_BAD_REQUEST)

        elif request.method == "DELETE":
            old_task = task.title
            # Only create logs for non-drafted tasks
            if not task.drafted:
                Logger.objects.create(
                    action=f"Task titled '{old_task}' has been removed from the system",
                    user=User.objects.get(id=task.employee.id)
                )
            task.delete()
            return Response(status=status.HTTP_202_ACCEPTED)

        elif request.method == "GET":
            serialized = TaskSerializer(task)
            return Response(serialized.data)

    except Exception as e:
        print(e)
        return Response(status=status.HTTP_418_IM_A_TEAPOT)


@api_view(['GET', 'POST'])
def company_admins(request):
    try:
        if request.method == "POST":
            request.data["password"] = make_password(request.data["password"])
            request.data["role"] = "company_admin"
            serialized = UserSerializer(data = request.data)
            if serialized.is_valid():
                serialized.save()
                return Response(serialized.data, status=status.HTTP_201_CREATED)

            else:
                return Response(status=status.HTTP_400_BAD_REQUEST)

        company_admins = User.objects.filter(role = "company_admin")
        serialized = UserSerializer(company_admins, many = True)
        return Response(serialized.data)
    except:
        return Response(status=status.HTTP_418_IM_A_TEAPOT)


@api_view(['GET', 'POST'])
def departments(request, pk):
    if request.method == "POST":
        request.data["company"] = pk
        serialized = DepartmentSerializer(data = request.data)
        if serialized.is_valid():
            serialized.save()
            serialized.data["admin_email"] = UserSerializer(User.objects.get(id = serialized.data["admin"])).data["email"]
            return Response(serialized.data, status=status.HTTP_201_CREATED)
        else:
            return Response(status=status.HTTP_400_BAD_REQUEST)

    departments = Department.objects.filter(company = pk)
    serialized = DepartmentSerializer(departments, many = True)
    for department in serialized.data:
            department["admin_email"] = UserSerializer(User.objects.get(id = department["admin"])).data["email"]
    return Response(serialized.data)


@api_view(['GET', 'POST'])
def department_admins(request, pk):
    try:
        if request.method == "POST":
            request.data["password"] = make_password(request.data["password"])
            request.data["role"] = "department_admin"
            request.data["company"] = pk
            serialized = UserSerializer(data = request.data)
            if serialized.is_valid():
                serialized.save()
                return Response(serialized.data, status=status.HTTP_201_CREATED)

            else:
                return Response(status=status.HTTP_400_BAD_REQUEST)

        department_admins = User.objects.filter(role = "department_admin", company = pk)
        serialized = UserSerializer(department_admins, many = True)
        return Response(serialized.data)
    except:
        return Response(status=status.HTTP_418_IM_A_TEAPOT)


@api_view(["GET", "POST"])
def employees(request, pk):
    try:
        if request.method == "POST":
            request.data["role"] = "employee"
            request.data["password"] = make_password(request.data["password"])
            request.data["department"] = pk
            serialized = UserSerializer(data = request.data)
            if serialized.is_valid():
                serialized.save()
                return Response(serialized.data, status = status.HTTP_201_CREATED)
            else:
                return Response(serialized.errors, status=status.HTTP_400_BAD_REQUEST)

        employees = User.objects.filter(role = "employee", department = pk)
        serialized = UserSerializer(employees, many =True)
        return Response(serialized.data)

    except:
        return Response(status=status.HTTP_418_IM_A_TEAPOT)


@api_view(["GET", "POST"])
def tasks(request, pk):
    try:
        if request.method == "POST":
            request.data["employee"] = pk
            serialized = TaskSerializer(data = request.data)
            if serialized.is_valid():
                serialized.save()
                return Response(serialized.data, status = status.HTTP_201_CREATED)
            else:
                return Response(serialized.errors, status=status.HTTP_400_BAD_REQUEST)

        tasks = Task.objects.filter(employee = pk, drafted = False)
        serialized = TaskSerializer(tasks, many =True)
        return Response(serialized.data)

    except:
        return Response(status=status.HTTP_418_IM_A_TEAPOT)


@api_view(['GET'])
def get_company(request, pk):
    try:
        company = Company.objects.get(admin = pk)
        serialized = CompanySerializer(company)
        return Response(serialized.data)
    except:
        return Response(status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
def get_department(request, pk):
    try:
        department = Department.objects.get(admin = pk)
        serialized = DepartmentSerializer(department)
        return Response(serialized.data)
    except:
        return Response(status=status.HTTP_400_BAD_REQUEST)

@api_view(['PATCH'])
def update_user(request, pk):
    try:
        user = User.objects.get(id = pk)
        try:
            if request.data["password"]:
                request.data["password"] = make_password(request.data["password"])
        except:
            pass
        serialized = UserSerializer(user, data = request.data, partial= True)
        if serialized.is_valid():
            serialized.save()
            return Response(serialized.data)
        else:
            return Response(status=status.HTTP_400_BAD_REQUEST)
    except:
        return Response(status = status.HTTP_418_IM_A_TEAPOT)


@api_view(['DELETE'])
def user(request, pk):
    try:
        user = User.objects.get(id = pk)
        user.delete()
        return Response(status = status.HTTP_202_ACCEPTED)
    except:
        return Response(status = status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'POST'])
def projects(request, pk):

    if request.method == "POST":
        request.data["company"] =pk
        serialized = ProjectSerializer(data = request.data)
        if serialized.is_valid():
            serialized.save()
            return Response(serialized.data,status = status.HTTP_201_CREATED)

    projects = Project.objects.filter(company = pk)
    serialized = ProjectSerializer(projects, many = True)
    return Response(serialized.data)


@api_view(['PATCH', 'DELETE'])
def project(request, pk):
    try:
        project = Project.objects.get(id = pk)
    except:
        project = None

    if project is None:
        return Response(status = status.HTTP_404_NOT_FOUND)

    if request.method == "DELETE":
        project.delete()
        return Response(status = status.HTTP_202_ACCEPTED)

    if request.method == "PATCH":
        serialized = ProjectSerializer(project, data = request.data, partial = True)
        if serialized.is_valid():
            serialized.save()
            return Response(serialized.data,status = status.HTTP_202_ACCEPTED)


@api_view(['GET'])
def task_grouping(request, pk):
    try:
        # today's details
        date_now = datetime.today()  # Fix the issue with datetime.date.today()
        year, week, _ = date_now.isocalendar()  # Extract year and week from today's date
        month = date_now.month  # Extract the current month
        day = date_now.day  # Extract today's day

        # Calculate previous month and year
        prev_month = month - 1 if month > 1 else 12
        prev_year = year if month > 1 else year - 1

        # get all employees in the company
        employees = User.objects.filter(
            Q(department__company__id=pk)
        )

        # store for results to be sent
        result = []

        # loop through all the employees while generating tasks
        for e in employees:
            # user instance
            user_data = {
                "id": e.id,
                "name": e.first_name + " " + e.last_name,
                "department": e.department.name,
                "company": e.department.company.id,
                "first_name": e.first_name,
                "contact": e.contact,
                "email": e.email
            }

            tasks = Task.objects.filter(employee=e.id, drafted=False)

            # statistics
            daily_tasks = {
                "task_count": 0,
                "total_duration": 0
            }
            weekly_tasks = {
                "task_count": 0,
                "total_duration": 0
            }
            monthly_tasks = {
                "task_count": 0,
                "total_duration": 0
            }
            previous_month_tasks = {
                "task_count": 0,
                "total_duration": 0
            }

            # loop over tasks for a given employee
            for task in tasks:
                # task date details
                task_date = task.created_at
                task_year, task_week, _ = task_date.isocalendar()  # Extract year and week
                task_month = task_date.month  # Extract the task's month
                task_day = task_date.day  # Extract the task's day

                if task_year == year:
                    # generate daily tasks
                    if task_day == day:
                        daily_tasks["task_count"] += 1
                        daily_tasks["total_duration"] += task.duration

                    # generate weekly tasks
                    if task_week == week:
                        weekly_tasks["task_count"] += 1
                        weekly_tasks["total_duration"] += task.duration

                    # generate monthly tasks
                    if task_month == month:
                        monthly_tasks["task_count"] += 1
                        monthly_tasks["total_duration"] += task.duration

                # generate previous month tasks
                if task_year == prev_year and task_month == prev_month:
                    previous_month_tasks["task_count"] += 1
                    previous_month_tasks["total_duration"] += task.duration

            # Add statistics to user data
            user_data["daily_tasks"] = daily_tasks
            user_data["monthly_tasks"] = monthly_tasks
            user_data["weekly_tasks"] = weekly_tasks
            user_data["previous_month_tasks"] = previous_month_tasks

            result.append(user_data)

        return Response(result)
    
    except Exception as e:
        # It's a good practice to log the exception for debugging purposes
        print(f"Error: {e}")
        return Response(status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
def task_grouping_by_department(request, pk):
    try:
        # today's details
        date_now = datetime.today()  # Use datetime.today() instead of datetime.date.today()
        year, week, _ = date_now.isocalendar()  # Extract year and week from today's date
        month = date_now.month  # Extract the current month
        day = date_now.day  # Extract today's day

        # Calculate previous month and year
        prev_month = month - 1 if month > 1 else 12
        prev_year = year if month > 1 else year - 1

        # get all employees in the department
        employees = User.objects.filter(
            Q(department__id=pk)
        )

        # store for results to be sent
        result = []

        # loop through all the employees while generating tasks
        for e in employees:
            # user instance
            user_data = {
                "id": e.id,
                "name": e.first_name + " " + e.last_name,
                "department": e.department.name,
                "company": e.department.company.id,
                "first_name": e.first_name,
                "contact": e.contact,
                "email": e.email
            }

            tasks = Task.objects.filter(employee=e.id, drafted=False)

            # statistics
            daily_tasks = {
                "task_count": 0,
                "total_duration": 0
            }
            weekly_tasks = {
                "task_count": 0,
                "total_duration": 0
            }
            monthly_tasks = {
                "task_count": 0,
                "total_duration": 0
            }
            previous_month_tasks = {
                "task_count": 0,
                "total_duration": 0
            }

            # loop over tasks for a given employee
            for task in tasks:
                # task date details
                task_date = task.created_at
                task_year, task_week, _ = task_date.isocalendar()  # Extract year and week
                task_month = task_date.month  # Extract the task's month
                task_day = task_date.day  # Extract the task's day

                if task_year == year:
                    # generate daily tasks
                    if task_day == day:
                        daily_tasks["task_count"] += 1
                        daily_tasks["total_duration"] += task.duration

                    # generate weekly tasks
                    if task_week == week:
                        weekly_tasks["task_count"] += 1
                        weekly_tasks["total_duration"] += task.duration

                    # generate monthly tasks
                    if task_month == month:
                        monthly_tasks["task_count"] += 1
                        monthly_tasks["total_duration"] += task.duration

                # generate previous month tasks
                if task_year == prev_year and task_month == prev_month:
                    previous_month_tasks["task_count"] += 1
                    previous_month_tasks["total_duration"] += task.duration

            # Add statistics to user data
            user_data["daily_tasks"] = daily_tasks
            user_data["monthly_tasks"] = monthly_tasks
            user_data["weekly_tasks"] = weekly_tasks
            user_data["previous_month_tasks"] = previous_month_tasks

            result.append(user_data)

        return Response(result)
    
    except Exception as e:
        # It's a good practice to log the exception for debugging purposes
        print(f"Error: {e}")
        return Response(status=status.HTTP_400_BAD_REQUEST)

# get allcompany members
@api_view(['GET', 'POST'])
def company_members(request, pk):

    # get all users whosecompany = pk, and role = company_member
    if request.method == "POST":
        request.data['username'] = f'{pk}.{request.data["email"]}'
        request.data["role"] = "company_member"
        request.data["password"] = make_password(request.data["password"])
        serialized = UserSerializer(data = request.data)
        if serialized.is_valid():
            serialized.save()
            return Response(serialized.data, status=status.HTTP_201_CREATED)

        else:
            return Response(status=status.HTTP_400_BAD_REQUEST)


    members = User.objects.filter(Q(username__startswith = str(pk) + ".") & Q(role ="company_member"))
    serialized = UserSerializer(members, many = True)
    return Response(serialized.data)

@api_view(["GET", "POST", "PATCH"])
def drafts(request, pk):
    try:
        tasks = Task.objects.filter(employee = pk, drafted = True)

        if request.method == "POST":
            request.data["employee"] = pk
            request.data["drafted"] = True
            serialized = TaskSerializer(data = request.data)
            if serialized.is_valid():
                serialized.save()
                return Response(serialized.data, status = status.HTTP_201_CREATED)
            else:
                return Response(serialized.errors, status=status.HTTP_400_BAD_REQUEST)

        elif request.method == "PATCH":
            tasks.update(drafted= False)
            return Response(status = status.HTTP_202_ACCEPTED)

        serialized = TaskSerializer(tasks, many =True)
        return Response(serialized.data)

    except:
        return Response(status=status.HTTP_418_IM_A_TEAPOT)


@api_view(['GET'])
def update_all_tasks(request):
    try:
        if Task.objects.filter(drafted = True).count() == 0:
            return Response({"msg": "Nothing to update"})

        Task.objects.all().update(drafted = False)
        return Response({"msg": "All pending drafts have been uploaded"}, status=status.HTTP_202_ACCEPTED)

    except:
        return Response(status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
def cron_job(request):
    try:
        Task.objects.create(
            title = "Testing cron jobs",
            body = "This is a demo",
            employee = User.objects.get(id = 6),
            duration = 10
        )
        return Response({"msg": "Task has been uploaded, in the next 2mins, another one will be uploaded"},status=status.HTTP_201_CREATED)
    except Exception as e:
        return Response({"msg": str(e)}, status=status.HTTP_400_BAD_REQUEST)


import os
import zipfile
from io import StringIO
from datetime import datetime
from django.core.management import call_command
from django.core.mail import EmailMessage
from django.conf import settings
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status

@api_view(['GET'])
def export_data_and_send_email(request):
    """
    Function-based view to dump database data into a JSON file, compress it,
    and send the compressed file via email with a timestamp in the subject.
    """
    buffer = StringIO()  # Create an in-memory buffer to capture the dumpdata output

    try:
        # Step 1: Call the dumpdata command to export the database data
        call_command('dumpdata', '--indent', '4', stdout=buffer)

        # Get the JSON data from the buffer
        json_data = buffer.getvalue()

        # Step 2: Save the JSON data to a file (TRS_DB.json)
        file_path = os.path.join(settings.BASE_DIR, 'TRS_DB.json')
        with open(file_path, 'w') as json_file:
            json_file.write(json_data)

        # Step 3: Compress the JSON file into a .zip file
        zip_file_path = os.path.join(settings.BASE_DIR, 'TRS_DB.zip')

        # Create a ZIP file and add the JSON file into it
        with zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            zipf.write(file_path, 'TRS_DB.json')

        # Step 4: Prepare the email with timestamped subject
        current_timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')  # Format: 2024-12-08_15-30-00
        subject = f'TRS_DB_{current_timestamp}.zip Export'
        body = f'Please find attached the compressed TRS_DB.zip file containing the exported database data as of {current_timestamp}.'
        to_email = 'kigongovincent625@gmail.com'

        # Create the email message
        email = EmailMessage(subject, body, settings.DEFAULT_FROM_EMAIL, [to_email])

        # Attach the compressed .zip file
        email.attach(f'TRS_DB_{current_timestamp}.zip', open(zip_file_path, 'rb').read(), 'application/zip')

        # Step 5: Send the email with the attachment
        email.send()

        # Return a success response
        return Response({'message': f'Data exported, compressed, and email sent successfully with timestamp: {current_timestamp}!'}, 
                        status=status.HTTP_200_OK)

    except Exception as e:
        # Handle any errors and return an error response
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    

# create logs     
@api_view(['GET'])
def logs(request, pk):
    logs = Logger.objects.filter(user__department__company__id=pk).order_by('-created_at')[:10]
    serialized = LoggerSerializer(logs, many=True)
    return Response(serialized.data)

@api_view(['GET'])
def department_monthly_report(request, pk):
    try:
        # Get current month and year
        current_date = datetime.today()
        current_month = current_date.month
        current_year = current_date.year

        # Get all employees in the department
        employees = User.objects.filter(
            department__id=pk,
            role='employee'
        )

        # Create a new workbook and select the active sheet
        wb = Workbook()
        ws = wb.active
        ws.title = f"Monthly Report {current_date.strftime('%B %Y')}"

        # Define styles
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        center_aligned = Alignment(horizontal='center', vertical='center')

        # Set column widths
        ws.column_dimensions['A'].width = 30  # Employee Name
        ws.column_dimensions['B'].width = 20  # Total Hours
        ws.column_dimensions['C'].width = 20  # Task Count

        # Add headers
        headers = ['Employee Name', 'Total Hours', 'Total Tasks']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_aligned

        # Add data for each employee
        row = 2
        for employee in employees:
            # Get tasks for current month
            monthly_tasks = Task.objects.filter(
                employee=employee,
                drafted=False,
                created_at__year=current_year,
                created_at__month=current_month
            )

            # Calculate totals
            total_hours = monthly_tasks.aggregate(total=Sum('duration'))['total'] or 0
            total_tasks = monthly_tasks.count()

            # Add employee data
            ws.cell(row=row, column=1, value=f"{employee.first_name} {employee.last_name}")
            ws.cell(row=row, column=2, value=total_hours)
            ws.cell(row=row, column=3, value=total_tasks)

            # Center align the data
            for col in range(1, 4):
                ws.cell(row=row, column=col).alignment = center_aligned

            row += 1

        # Add department info at the top
        department = Department.objects.get(id=pk)
        ws.insert_rows(1, 2)
        ws.merge_cells('A1:C1')
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = f"{department.name} - Monthly Report"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Create the response
        response = HttpResponse(
            content_type='application/vnd.openpyxl-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=department_report_{current_date.strftime("%Y_%m")}.xlsx'

        # Save the workbook to the response
        wb.save(response)
        return response

    except Exception as e:
        print(f"Error generating report: {e}")
        return Response(
            {'error': 'Failed to generate report'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
